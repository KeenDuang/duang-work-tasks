# -*- coding: utf-8 -*-
"""
Created on Fri Oct 16 16:55:01 2020

@author: kedongwang
"""

import time
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from os import remove, path

print('开始处理')
start = time.time()

"""excel格式调整类"""


class XlsxSaver:
    """
    一个将DataFrame转换成格式化excel的工具
    """

    def __init__(self, df_in, filename='a.xlsx', sheet_name='Sheet1'):
        """
        df_in : 从一个DataFrame对象获取表格内容
        filename : 文件名
        sheet_name : 表名
        """
        self.filename = filename  # 保存的xlsx文件的名字
        self.user_def = []  # 储存由用户自定义的列的列名，这些列不再参与自动计算列宽
        if path.exists(filename):
            # 如果文件存在，就直接打开，添加Sheet
            self.wb = load_workbook(filename)
            self.sheet = self.wb.create_sheet(sheet_name)
        else:
            # 如果文件不存在，就创建表格
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.title = sheet_name
        # 将df的内容复制给sheet
        self.df = df_in.copy()
        self.sheet.append(list(self.df.columns))
        for row in range(0, len(list(self.df.index))):
            for col in range(0, len(list(self.df.columns))):
                self.sheet.cell(row + 2, col + 1).value = self.df.iloc[row, col]  # 注意：sheet行列从1开始计数

    def remove_file(self):
        remove(self.filename)

    def set_sheet_name(self, sheet_name):
        self.sheet.title = sheet_name

    def set_filename(self, filename):
        self.filename = filename

    def get_maxlength(self, series_in, col):
        """
        获取一个类型为object的Series中的最大占位长度，用于确定导出的xlsx文件的列宽
        col : 表头，也参与比较，解决有时候表头过长的问题
        """
        series = series_in.fillna('-')  # 填充空值，防止出现nan
        str_list = list(series)
        len_list = []
        for elem in str_list + [col]:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1
                else:
                    length += 2
            len_list.append(length)
        return max(len_list)

    def __auto_width(self):
        cols_list = list(self.df.columns)  # 获取列名
        for i in range(0, len(cols_list)):
            col = cols_list[i]
            if col in self.user_def:
                continue
            self.sheet.cell(1, i + 1).font = Font(bold=True)  # 加粗表头
            letter = chr(i + 65)  # 由ASCII值获得对应的列字母
            max_len = self.get_maxlength(self.df[col].astype(str), col)
            # if max_len <= 12:
            #    self.sheet.column_dimensions[letter].width = 14
            if max_len <= 50:
                self.sheet.column_dimensions[letter].width = max_len + 2
                for cell in self.sheet[letter]:
                    cell.alignment = Alignment(wrap_text=True, horizontal='left')
            else:
                self.sheet.column_dimensions[letter].width = 50
                for cell in self.sheet[letter]:
                    cell.alignment = Alignment(wrap_text=True, horizontal='left')

    def save(self):
        # 自动调整列宽，并保存
        self.__auto_width()
        self.wb.save(self.filename)


"""账户信息比对代码"""

para = []
with open("../data/机构参数.txt", "r", encoding='utf-8') as f:
    for line in f.readlines():
        line = line.strip('\n')
        para.append(line)

# ##  核心与人行数据对比

# # 数据读取
hexin = pd.read_excel('../data/核心.xls')
renhang = pd.read_excel('../data/人行.xls')

# #  条件筛选
hexin = hexin[(hexin['账户属性'] == '4 - 基本户') |
              (hexin['账户属性'] == '3 - 一般户') |
              (hexin['账户属性'] == '6 - 专用户（核准类）') |
              (hexin['账户属性'] == '23 - 专用户（备案类）') |
              (hexin['账户属性'] == '5 - 临时户')]
hexin.rename(columns={'账号/卡号': '账号'}, inplace=True)

# # 重复值处理
hexin['销户日期'].fillna(20991231, inplace=True)
hexin.sort_values(by=['开户日期', '销户日期'], ascending=False, inplace=True)
hexin.drop_duplicates(subset=['账号'], keep='first', inplace=True)
hexin = hexin.sort_values(by=['账号']).reset_index(drop=True)

renhang['销户日期'].fillna('2099-12-31', inplace=True)
renhang.sort_values(by=['开户日期', '销户日期'], ascending=False, inplace=True)
renhang.drop_duplicates(subset=['账号'], keep='first', inplace=True)
renhang = renhang.sort_values(by=['账号']).reset_index(drop=True)

# # 账号异常值处理
hexin = hexin[hexin['账号'].map(lambda x: len(str(x))) >= 6].reset_index(drop=True)
renhang = renhang[renhang['账号'].map(lambda x: len(str(x))) >= 6].reset_index(drop=True)

# # 核心无人行有
hexin['账号'] = hexin['账号'].map(lambda x: str(x))
res1 = pd.DataFrame(list(set(renhang['账号']) - set(hexin['账号'])), columns=['人行账号'])
columns = ['人行账号', '账户名称', '账户性质', '开户日期', '状态']
res1 = res1.merge(renhang, how='inner', left_on='人行账号', right_on='账号')[columns]
res1.columns = ['人行账号', '人行账户名称', '人行账户性质', '人行开户日期', '人行账户状态']
res1 = res1[res1['人行账户状态'] != '撤销']
res1['备注'] = None
res1.insert(0, column='机构名称', value=para[1].split('：')[1])
res1.insert(0, column='机构号', value=para[0].split('：')[1])

# # 核心有人行无
res2 = pd.DataFrame(list(set(hexin['账号']) - set(renhang['账号'])), columns=['核心账号'])
columns = ['核心账号', '户名', '账户类型', '账户属性', '开户日期', '账户状态']
res2 = res2.merge(hexin, how='inner', left_on='核心账号', right_on='账号')[columns]
res2.columns = ['核心账号', '核心户名', '核心账户类型', '核心账户属性', '核心开户日期', '核心账户状态']
res2 = res2[res2['核心账户状态'] != 'C-销户']
res2['备注'] = None
res2.insert(0, column='机构名称', value=para[1].split('：')[1])
res2.insert(0, column='机构号', value=para[0].split('：')[1])

# # 核心与人行户名不一致
res3 = hexin.merge(renhang, how='inner', on='账号')
res3 = res3[res3['户名'] != res3['账户名称']]
res3 = res3[['账号', '户名', '账户名称', '账户类型', '账户属性', '账户性质', '开户日期_x', '开户日期_y', '账户状态', '状态']]
res3.columns = ['账号', '核心户名', '人行账户名称', '核心账户类型', '核心账户属性', '人行账户性质', '核心开户日期',
                '人行开户日期', '核心账户状态', '人行状态']


def pre3(x):
    x = x.replace('（', '')
    x = x.replace('）', '')
    x = x.replace('(', '')
    x = x.replace(')', '')
    x = x.replace('０', '0')
    x = x.replace('１', '1')
    x = x.replace('２', '2')
    x = x.replace('３', '3')
    x = x.replace('４', '4')
    x = x.replace('５', '5')
    x = x.replace('６', '6')
    x = x.replace('７', '7')
    x = x.replace('８', '8')
    x = x.replace('９', '9')
    return x


res3['核心户名_'] = res3['核心户名'].map(lambda x: pre3(x))
res3['人行账户名称_'] = res3['人行账户名称'].map(lambda x: pre3(x))
res3 = res3[res3['核心户名_'] != res3['人行账户名称_']]
res3 = res3[res3['核心账户状态'] != 'C-销户']
res3['备注'] = None
res3.insert(0, column='机构名称', value=para[1].split('：')[1])
res3.insert(0, column='机构号', value=para[0].split('：')[1])
res3.drop(columns=['核心户名_', '人行账户名称_'], inplace=True)

# # 核心与人行账号属性不一致
res4 = hexin.merge(renhang, how='inner', on='账号')


def pre4(x):
    if '基本' in x: return '基本'
    if '一般' in x: return '一般'
    if '专用' in x: return '专用'
    if '临时' in x: return '临时'
    if '验资' in x: return '验资'


res4['账户属性_'] = res4['账户属性'].apply(lambda x: pre4(x))
res4['账户性质_'] = res4['账户性质'].apply(lambda x: pre4(x))
res4 = res4[res4['账户属性_'] != res4['账户性质_']]
res4 = res4[['账号', '户名', '账户名称', '账户类型', '账户属性', '账户性质', '开户日期_x', '开户日期_y', '账户状态', '状态']]
res4.columns = ['账号', '核心户名', '人行账户名称', '核心账户类型', '核心账户属性', '人行账户性质', '核心开户日期',
                '人行开户日期', '核心账户状态', '人行状态']
res4 = res4[res4['核心账户状态'] != 'C-销户']
res4['备注'] = None
res4.insert(0, column='机构名称', value=para[1].split('：')[1])
res4.insert(0, column='机构号', value=para[0].split('：')[1])

# #### 核心与人行账号状态不一致

res5 = hexin.merge(renhang, how='inner', on='账号')


def pre5(x):
    if 'C-销户' in x:
        return '撤销'
    if ('D-不动户' in x) or ('A-活动' in x) or ('N-新建' in x) or ('H-预开户' in x):
        return '正常'
    if 'S-久悬户' in x:
        return '久悬'


res5['账户状态_'] = res5['账户状态'].apply(lambda x: pre5(x))
res5 = res5[res5['账户状态_'] != res5['状态']]
res5 = res5[['账号', '户名', '账户名称', '账户类型', '账户属性', '账户性质', '开户日期_x', '开户日期_y', '账户状态', '状态']]
res5.columns = ['账号', '核心户名', '人行账户名称', '核心账户类型', '核心账户属性', '人行账户性质', '核心开户日期',
                '人行开户日期', '核心账户状态', '人行状态']
res5['备注'] = None
res5.insert(0, column='机构名称', value=para[1].split('：')[1])
res5.insert(0, column='机构号', value=para[0].split('：')[1])

# ##  核心与验印数据对比

# # 数据读取
hexin = pd.read_excel('../data/核心.xls')
yanyin = pd.read_excel('../data/验印.xls')

# #  条件筛选
hexin.rename(columns={'账号/卡号': '账号'}, inplace=True)
yanyin = yanyin[(yanyin['账户属性'] == '基本户') |
                (yanyin['账户属性'] == '一般户') |
                (yanyin['账户属性'] == '专用户（核准类）') |
                (yanyin['账户属性'] == '专用户（备案类）') |
                (yanyin['账户属性'] == '临时户') |
                (yanyin['账户属性'] == '验资户')]
yanyin = yanyin[yanyin['销户标志'] == '正常']
yanyin = yanyin[yanyin['主从账户标志'] != '从账户']

# # 重复值处理
hexin['销户日期'].fillna(20991231, inplace=True)
hexin.sort_values(by=['开户日期', '销户日期'], ascending=False, inplace=True)
hexin.drop_duplicates(subset=['账号'], keep='first', inplace=True)
hexin = hexin.sort_values(by=['账号']).reset_index(drop=True)

yanyin['销户日期'].fillna('2099-12-31', inplace=True)
yanyin.sort_values(by=['开户日期', '销户日期'], ascending=False, inplace=True)
yanyin.drop_duplicates(subset=['账号'], keep='first', inplace=True)
yanyin = yanyin.sort_values(by=['账号']).reset_index(drop=True)

# # 账号异常值处理
hexin = hexin[hexin['账号'].map(lambda x: len(str(x))) >= 6].reset_index(drop=True)
yanyin = yanyin[yanyin['账号'].map(lambda x: len(str(x))) >= 6].reset_index(drop=True)

# # 核心有验印无
hexin['账号'] = hexin['账号'].map(lambda x: str(x))
res6 = pd.DataFrame(list(set(hexin['账号']) - set(yanyin['账号'])), columns=['账号'])
columns = ['账号', '户名', '账户类型', '账户属性', '开户日期', '账户状态']
res6 = res6.merge(hexin, how='inner', on='账号')[columns]
res6.columns = ['核心账号', '核心户名', '核心账户类型', '核心账户属性', '核心开户日期', '核心账户状态']
res6 = res6[res6['核心账户状态'] != 'C-销户']
res6['备注'] = None
res6.insert(0, column='机构名称', value=para[1].split('：')[1])
res6.insert(0, column='机构号', value=para[0].split('：')[1])

# # 核心无验印有
res7 = pd.DataFrame(list(set(yanyin['账号']) - set(hexin['账号'])), columns=['账号'])
columns = ['账号', '账户名称', '账户属性', '开户日期']
res7 = res7.merge(yanyin, how='inner', on='账号')[columns]
res7.columns = ['验印账号', '验印账户名称', '验印账户属性', '验印开户日期']
res7['备注'] = None
res7.insert(0, column='机构名称', value=para[1].split('：')[1])
res7.insert(0, column='机构号', value=para[0].split('：')[1])

# # 核心与验印账户属性不一致
res8 = hexin.merge(yanyin, how='inner', on='账号')
res8['账户属性_x_'] = res8['账户属性_x'].apply(lambda x: pre4(x))
res8['账户性质_y_'] = res8['账户属性_y'].apply(lambda x: pre4(x))
res8 = res8[res8['账户属性_x_'] != res8['账户性质_y_']]
res8 = res8[['账号', '户名', '账户名称', '账户类型', '账户属性_x', '账户属性_y', '开户日期_x', '开户日期_y', '账户状态']]
res8.columns = ['账号', '核心户名', '验印账户名称', '核心账户类型', '核心账户属性', '验印账户属性', '核心开户日期', '验印开户日期', '核心状态']
res8 = res8[res8['核心状态'] != 'C-销户']
res8['备注'] = None
res8.insert(0, column='机构名称', value=para[1].split('：')[1])
res8.insert(0, column='机构号', value=para[0].split('：')[1])

# # 核心与验印户名不一致
res9 = hexin.merge(yanyin, how='inner', on='账号')
res9 = res9[res9['户名'] != res9['账户名称']]
res9 = res9[res9['印鉴状态'] != '已抽卡']
res9 = res9[['账号', '户名', '账户名称', '账户类型', '账户属性_x', '账户属性_y', '开户日期_x', '开户日期_y', '账户状态']]
res9.columns = ['账号', '核心户名', '验印账户名称', '核心账户类型', '核心账户属性', '验印账户属性', '核心开户日期', '验印开户日期', '核心状态']
res9['核心户名_'] = res9['核心户名'].map(lambda x: pre3(x))
res9['验印账户名称_'] = res9['验印账户名称'].map(lambda x: pre3(x))
res9 = res9[res9['核心户名_'] != res9['验印账户名称_']]
res9 = res9[res9['核心状态'] != 'C-销户']
res9['备注'] = None
res9.insert(0, column='机构名称', value=para[1].split('：')[1])
res9.insert(0, column='机构号', value=para[0].split('：')[1])
res9.drop(columns=['核心户名_', '验印账户名称_'], inplace=True)

# ### 导出

XlsxSaver(res1, '../result/分析结果.xlsx', '核心无人行有').save()
XlsxSaver(res2, '../result/分析结果.xlsx', '核心有人行无').save()
XlsxSaver(res3, '../result/分析结果.xlsx', '核心与人行户名不一致').save()
XlsxSaver(res4, '../result/分析结果.xlsx', '核心与人行账号属性不一致').save()
XlsxSaver(res5, '../result/分析结果.xlsx', '核心与人行账号状态不一致').save()
XlsxSaver(res6, '../result/分析结果.xlsx', '核心有验印无').save()
XlsxSaver(res7, '../result/分析结果.xlsx', '核心无验印有').save()
XlsxSaver(res8, '../result/分析结果.xlsx', '核心与验印账户属性不一致').save()
XlsxSaver(res9, '../result/分析结果.xlsx', '核心与验印户名不一致').save()

end = time.time()
print("分析完成，用时%d秒" % round(end - start))
